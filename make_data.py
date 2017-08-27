import openpyxl
import csv

#읽어올 Execl파일명,Sheet,행과열을  불러와 수식을 이용 변환
def open_execl(execl_name):
    wb = openpyxl.load_workbook(execl_name)
    ws = wb.get_sheet_by_name("Sheet1")
    result = []
    for col in ws.iter_cols(min_row=1, min_col=3, max_col=3, max_row=3):
        for cell in col:
            a = str(cell.value)
            hour_a = int(a[0:2]) # 시간을 분으로 계산
            minute_a = int(a[3:5])
            delay_time_a = hour_a * 60 + minute_a
            result.append(delay_time_a)

#엑셀의 C5번 칸의 값을 불러와 분으로 변환 후 결과값에 추가
    b = str(ws['C5'].value)
    hour_b = int(b[0:2])
    minute_b = int(b[3:5])
    delay_time_b = hour_b * 60 + minute_b
    result.append(delay_time_b)
    return(result)

#계산한 값을 csv파일로 저장
def save_csv(file_name, result):
    f = open(file_name,'a',newline='')
    cswrite = csv.writer(f)
    cswrite.writerow(result)
    f.close()


def main():

    csv_name = str(input("저장할 csv파일명을 입력하시요:"))+".csv"

    while True:
        execl_name = str(input("저장할 xlsx파일명을 입력하시요(없으면 n을 입력):"))+".xlsx"

        if execl_name == "n.xlsx":
            print("프로그램이 끝났습니다.")
            break
        else:
            result = open_execl(execl_name)
            save_csv(csv_name, result)




if __name__ == "__main__":
    main()

import openpyxl
wb = openpyxl.load_workbook('aaa.xlsx')
#ws = wb.active
ws = wb.get_sheet_by_name("Sheet1")

#d = ws.cell(row=4, column=2).value
#print(ws['A2'].value)
#print(d)


#for r in ws.rows:
#    A = r[0].value
#    print(A)

#for col in ws.iter_cols(min_row=1, max_col=1, max_row=2):
#    for cell in col:
#         print(cell.value)


#for row in ws.iter_rows(min_row=1, max_col=2, max_row=1):
#    for result in row:
#         print(result.value)

#a=ws['A1'].value
#b=ws['A2'].value
#c=str(a)+str(b)
#print(c)

#------------------------------------------------------------------
#a=str(ws['A1'].value)

#print(a)
#Hour = int(a[0:2])
#Minute = int(a[3:5])
#delay_time = Hour * 60 + Minute

#print(delay_time)

#txt.파일로 쓰기---------------------------------------------------------
#f = open("aaa.txt",'w')
#for col in ws.iter_cols(min_row=1, min_col=3, max_col=3, max_row=5):
#    for cell in col:
#        a = str(cell.value)
#        Hour = int(a[0:2])
#        Minute = int(a[3:5])
#        delay_time = Hour * 60 + Minute

#        data = "%d " % delay_time
#        f.write(data)
#f.close

#CSV 파일로 쓰기-----------------------------------------------------------
import csv

f = open('output.csv', 'w', encoding='utf-8', newline="")
for col in ws.iter_cols(min_row=1, min_col=3, max_col=3, max_row=5):
    for cell in col:
        a = str(cell.value)
        Hour = int(a[0:2])
        Minute = int(a[3:5])
        delay_time = Hour * 60 + Minute

        wr = csv.writer(f)
        wr.writerow([delay_time])

f.close
